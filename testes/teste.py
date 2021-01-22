from openpyxl import  load_workbook
import requests

wb = load_workbook('ev.xlsx')
ws = wb['Acervo']

max_cols = ws.max_column
max_rows = ws.max_row


data = {
    "login": "gmendes",
    "password": "teste123",
    "firstname": "Guilherme",
    "lastname": "Monteiro"
}

credentials = {
    "login": data["login"],
    "password": data["password"]
}

auth = requests.post('http://127.0.0.1:5000/authorization', data=credentials)
print(auth.status_code)

if auth.status_code == 401:    
    response = requests.post(url='http://127.0.0.1:5000/users/register', data=data)
    auth = requests.post('http://127.0.0.1:5000/authorization', data=credentials)

access_token = auth.json()['access_token']
headers={'Authorization': f'Bearer {access_token}' }
print(headers)

for i in range(2, max_rows):

    
    data = {
        "uuid_ev": ws.cell(row=i, column=1).value,
        "erro": ws.cell(row=i, column=2).value,
        "isbn_issn": ws.cell(row=i, column=3).value,
        "autor": ws.cell(row=i, column=4).value,
        "titulo": ws.cell(row=i, column=5).value,
        "editora": ws.cell(row=i, column=6).value,
        "ano": ws.cell(row=i, column=7).value,
        "estante": ws.cell(row=i, column=8).value,
        "preco": ws.cell(row=i, column=9).value,
        "conservacao": ws.cell(row=i, column=10).value,
        "peso": ws.cell(row=i, column=11).value,
        "tipo_publicacao": ws.cell(row=i, column=12).value,
        "tipo_produto": ws.cell(row=i, column=13).value,
        "edicao_numero": ws.cell(row=i, column=14).value,
        "numero": ws.cell(row=i, column=15).value,
        "idioma": ws.cell(row=i, column=16).value,
        "volume": ws.cell(row=i, column=17).value,
        "acabamento": ws.cell(row=i, column=18).value,
        "desconto": ws.cell(row=i, column=19).value,
        "marcas_de_uso": ws.cell(row=i, column=20).value,
        "amarelo": ws.cell(row=i, column=21).value,
        "manchas_bolor": ws.cell(row=i, column=22).value,
        "grifos_anotações": ws.cell(row=i, column=23).value,
        "amassados_orelha": ws.cell(row=i, column=24).value,
        "rasgos": ws.cell(row=i, column=25).value,
        "dedicatoria": ws.cell(row=i, column=26).value,
        "autografo": ws.cell(row=i, column=27).value,
        "numeric_id": ws.cell(row=i, column=28).value,
        "assunto": ws.cell(row=i, column=29).value,
        "localizacao": ws.cell(row=i, column=30).value,
        "seller": 'sebolinhapaulista'
    }

    post_edtion = requests.post(url='http://127.0.0.1:5000/estantevirtual', data=data, headers=headers)
    print(post_edtion.status_code)
    print(post_edtion.json())
    
    